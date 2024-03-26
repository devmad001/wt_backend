import re
import sys
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from shutil import copyfile


#0v8#  JC Sep 11, 2023  WT
#0v7#  JC Feb 15, 2023  pharmabrain
#0v6#  JC Jul 26, 2020  Use with ambassador
#                       - Style no longer used
#                       - py3
#0v5#  JC Aug  8, 2018  via ash 18 8 6
#0v4#  JC June 1, 2018  Updated
#0v3#  JC May 25, 2017  Inhereted
#0v2#  JC Jan 4, 2017   Inheret from e5_transpose_2.py


class NestedDict(dict):
    def __getitem__(self, key):
        if key in self: return self.get(key)
        return self.setdefault(key, NestedDict())


class Excel_Wrap(object):
    # - may be extra extending logic in ie/ e5_transpose_2_D1x...
    def __init__(self):
        self.filename=''
        #unused#  self.smem=NestedDict()
        self.active_sheet=''
        return
    
    def create_copy_from_to(self,from_filename,to_filename):
        copyfile(from_filename,to_filename)
        return to_filename
    
    def open(self,filename):
        self.filename=filename
        self.wb = load_workbook(self.filename)

        try:
            pass
            #self.wb.save(self.filename)
        except:
            print ("Please close or ensure file exists: "+str(self.filename))
            a=please_close_excel_file
        return

    def save(self):
        self.wb.save(self.filename)
        return self.filename

    def load_sheet(self,sheet_name=''):
        print (str(self.wb.get_sheet_names()))
        if not sheet_name in self.wb.get_sheet_names():
            sheet_name=self.wb.get_sheet_names()[0]
        sheet=self.wb[sheet_name]
        self.active_sheet=sheet
        return sheet
    
    def iter_rows_dict(self,sheet='',min_header_length=3,header_row=0):
        # Recover header and yiel dicts
        headers={}

        if not sheet:
            sheet=self.active_sheet
        if isinstance(sheet,str):
            sheet=self.wb[sheet] #patch
        
        row_num=0
        for row in sheet.iter_rows():
            row_num+=1
            row_line=[]
            none_non=False
            
            if header_row and row_num<header_row:
                print ("[debug] skipping row #"+str(row_num)+" cause waiting for header line")
                continue

            for cell in row:
                row_line.append(cell.value)
                if cell.value: none_non=True

            if not none_non: row_line=[]  #All None is blank row

            if not headers and len(row_line)>min_header_length:
#D#                print ("[debug] assume header: "+str(row_line))
                j=-1
                for header in row_line:
                    j+=1
                    headers[j]=header
                print ("[debug] using header columns: "+str(headers))
            elif headers:
                #transform line into dict
                if row_line:
                    dd={}
                    j=-1
                    for item in row_line:
                        j+=1
                        dd[headers[j]]=item
                    yield dd
        return
    
    def iter_rows_custom(self,sheet=''):
        if not sheet:
            sheet=self.active_sheet
        if isinstance(sheet,str):
            sheet=self.wb[sheet] #patch
        for row in sheet.iter_rows():
            row_line=[]
            for cell in row:
                row_line.append(cell.value)
            yield row_line
        return

    def set(self,row,col,value):
        self.active_sheet.cell(row=row,column=col).value=str(value)
        return
    
    def get_headers(self,sheet,the_min,the_max,mapping={}):
        #if not the_min==1:a=stop_assume_name_matches_numbers_for_now
        #For now assume range
        cc=the_min
        headers=[]
        while(cc<the_max):
            cc+=1
            colname=sheet.cell(row=1,column=cc).value
            colname=mapping.get(colname,colname) #Mapping if necessary
            headers.append(colname)
        return headers
    
    def check_max_limits(self,sheet): #dd
        new_max_col=sheet.max_column
        new_max_row=sheet.max_row
        print ("Determining max workspace size... given col:"+str(new_max_col)+", row: "+str(new_max_row))

        factor=1
        found=False
        while(not found):
            while (new_max_col>0):
                if sheet.cell(row=1,column=new_max_col).value:
                    break
                else: new_max_col-=1

            if sheet.max_column==new_max_col:
                factor=factor*1.20
                new_max_col=int(sheet.max_column*factor)
            else:found=True
        factor=1
        found=False
        while(not found):
            while (new_max_row>0):
                if sheet.cell(row=new_max_row,column=1).value:
                    break
                else: new_max_row-=1

            if sheet.max_row==new_max_row:
                factor=factor*1.20
                new_max_row=int(sheet.max_row*factor)
            else:found=True
                
        self.max_row=new_max_row
        self.max_col=new_max_col
        print ("Active worksheet size: "+str(self.max_col)+" x "+str(self.max_row)+"= "+str(self.max_col*self.max_row))
        return
    
    def iter_row(self,sheet=''):
        rowp=1 #1 ignores header
        while (rowp<self.max_row):
            rowp+=1
            yield rowp
        return
    
    def get_row(self,sheet,idx):
        c=1
        row=[]
        while c<=self.max_col:
            row+=[sheet.cell(row=idx,column=c).value]
            c+=1
        return row

    #Above is validated
    #########################################
    

    def old_load_sheet(self,sheet_name=''):
        #UNUSED
        #    for row_line in ws.iter_rows():
        #    for cell in row_line:
        #        cell.value = ""
        #*unused
        #        if sheet_name in self.wb2.get_sheet_names:
        #self.sheet_name=sheet_name
        #max_row=1
        #blanks=0
        #sheet=self.wb[self.sheet_name]
        #for row in range(1,max_row):
        #    for col in range(1,self.max_col):
        #        value1=sheet.cell(row=row,column=col).value
        #        #value=sheet.cell(self.title_col+str(row)).value
        #        if value1 is not None:
        #            pass
        #            #unused#  self.smem[row][col]=value1
        #            #print str(row)+","+str(col)+": "+str(value1)
        #return
        return
    
    def row2row(self,target_sheet,row_list,row_num):
        #Supports between two sheets
        c=1
        for col_value in row_list:
            target_sheet.cell(row=row_num,column=c).value=col_value
            c+=1
        return

    def copy_row(self,sheet_from,sheet_to,row_from,row_to):
        sheet_from=self.wb[sheet_from]
        sheet_to=self.wb[sheet_to]

        col_from=0
        while (col_from<self.max_col):
            col_from+=1
            sheet_to.cell(row=row_to,column=col_from).value=sheet_from.cell(row=row_from,column=col_from).value
        return

    def del_row(self,sheet,row):
        sheet_from=self.wb[sheet]
        col_from=0
        while (col_from<self.max_col):
            col_from+=1
            sheet_from.cell(row=row,column=col_from).value=''
        return

    def move_row(self,sheet_from,sheet_to,row_from,row_to):
        self.copy_row(sheet_from,sheet_to,row_from,row_to)
        self.del_row(sheet_from,row_from)
        return


    
    def del_col_contents(self,s, col_num, cmax=None, rmax=None,gap=1): #d1 def
        start_row=2 #Leave header
        self.max_row=self.max_row #reminder
        print ("[debug] removing rows upto row: "+str(self.max_row+1))
        
        for r in range(start_row, self.max_row+1):
            s.cell(row=r,column=col_num).value=''
        return s

    def del_col(self,s, col_num, cmax=None, rmax=None,gap=1): #d1 def
        #http://stackoverflow.com/questions/22048016/delete-column-from-xls-file
        #Gap if expecting jump
        cc=(cmax-gap-col_num)*rmax
        ccc=cmax-gap
        print ("cmax: "+str(cmax))
        print ("gap: "+str(gap))
        print ("col_num: "+str(col_num))

        print ("[debug moving cell count]: "+str(cc))
        for c in range(col_num, cmax - 1):
            print ("Deleting at column "+str(c)+" of "+str(cmax))
            for r in range(1, rmax):
                s.cell(row=r,column=c).value=s.cell(row=r,column=c+gap).value
                s.cell(row=r,column=c+gap).value=''
        print ("Done del col")
        return s
    
    def colname_index(self,sheet,colname):
        found=False
        #Given column name, return its' index
        for c in range(1, self.max_col+1):
            tt=sheet.cell(row=1,column=c).value
            if colname==tt:
                found=True
#D                print "Found: "+str(tt)
                break
            else:
                #print "Nope: "+str(tt)
                pass
        if not found:
            print ("Not found column last one tried was: "+str(tt))
            c+=1 #assume end plus 1
        return c
    

def test_copy_row():
    filename="test1.xlsx"
    Excel=Excel_Wrap()
    print ("Opening "+str(filename)+"...")
    Excel.open(filename)
    
    sheet_from='Sheet1'
    sheet_to='Sheet2'
    row_from=1
    row_to=1
    
    sheet_from=Excel.wb[sheet_from]
    sheet_to=Excel.wb[sheet_to]
    Excel.check_max_limits(sheet_from)
    
    print ("Ok...")
    
    from openpyxl.utils import get_column_letter
    min_col=1
    max_col=Excel.max_col
    min_row=row_from
    max_row=row_from
    rows=sheet_from.get_squared_range(min_col, min_row, max_col, max_row)
    row_vals=[]
    for row in rows:
        for cell in row:
            row_vals.append(cell.value)
    print ("GOT: "+str(row_vals))

    min_col_letter=get_column_letter(min_col)
    max_col_letter=get_column_letter(max_col)
    sheet_to[min_col_letter+str(min_row):max_col_letter+str(max_row)].value=row_vals
    #NOTE:  target value must still be set cell by cell
    #https://stackoverflow.com/questions/37632327/openpyxl-assign-value-or-apply-format-to-a-range-of-excel-cells-without-iterati
    a=ok
            
#    for row in sheet_from.iter_rows(min_row=row_from, max_col=Excel.max_col, max_row=row_from):
#        print "GOT: "+str(row)
#    a=kk
#...    for cell in row:
#...        print(cell)


#    col_from=0
#    while (col_from<Excel.max_col):
#        col_from+=1
#        sheet_to.cell(row=row_to,column=col_from).value=sheet_from.cell(row=row_from,column=col_from).value

#    cellrange = sheet1['A1':'B20'].value
#    sheet2['A1':'B20'] = cellrange
        
        
    Excel.save()
    
    
    return

if __name__ == '__main__':            
    test_copy_row()
    
    
    
    
    
    
    
