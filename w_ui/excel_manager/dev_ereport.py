import os
import sys
import codecs
import json
import re

LOCAL_PATH = os.path.abspath(os.path.dirname(__file__))+"/"
sys.path.insert(0,LOCAL_PATH)
sys.path.insert(0,LOCAL_PATH+"../")

from openpyxl.styles import Font  #external import but fine for style

from exceltemplate.excel_wrap import Excel_Wrap
#from exceltemplate.excel_markup import do_markup

## Tunnel to data source [ ] use api
sys.path.insert(0,LOCAL_PATH+"../../") 
from a_query.std_query import dev_query_get_excel_data_raw_1 

#from v_questions.excel_question import answer_excel_1
from v_questions.excel_question_v2 import answer_excel_2 as answer_excel


#0v3# JC  Jan  4, 2024  New opening balance option
#0v2# JC  Oct 14, 2023  Integrate with fresh data pull (/v_questions/excel_questeion.py)
#0v1# JC  Sep 11, 2023  Inheret from pharma


"""
    Markup Excel template with data
    - Excel_Wrap is base library
    - excel_markup is sample markup
"""

def iter_case_source_data(case_id,target_cols=[]):
    ## Include map
    #[ ] check target==source

    result_dicts,meta=dev_query_get_excel_data_raw_1(case_id)

    for result in result_dicts:
        ## MAPPINGS
        # amount,description,date
        dd={}
        dd['Acct #']='xxxx'
        dd['Eff Date']=result.get('date','')
        dd['Posted Date']='xxx'
        dd['Description']=result.get('description','')
        dd['Hidden']=''
        dd['Check #']=''
        dd['Debit']=result.get('amount','')
        dd['Credit']=result.get('amount','')
        dd['Balance']=''

        yield dd,target_cols

    return

def iter_source_data(target_cols=[]):
    ## Include map
    #[ ] check target==source

    ## Assign random data
    record_count=10
    for i in range(record_count):
        dd={}
        dd['Acct #']='1234'
        dd['Eff Date']='13/12/2018'
        dd['Posted Date']=''
        dd['Description']='Demo #'+str(i)
        dd['Hidden']=''
        dd['Check #']=''
        dd['Debit']='0'
        dd['Credit']='0'
        dd['Balance']=''
        yield dd,target_cols

    return

def dev_generate_excel_report_1(case_id='SGM BOA'):
    version="1"
    case_filename='excel_case-'+str(case_id)+"-"+str(version)+'.xlsx'

    print ("[info] running excel for case: "+str(case_id))

    Excel=Excel_Wrap()

    excel_dir=LOCAL_PATH+"excel_files/"
    excel_case_dir=excel_dir+case_id+"/"
    # Create dir if not exists
    if not os.path.exists(excel_case_dir):
        os.makedirs(excel_case_dir)
    template_filename=excel_dir+"raw_template1.xlsx"
    target_filename=excel_case_dir+case_filename

    Excel.create_copy_from_to(template_filename,target_filename)

    print ("1)  Load template")

    Excel.open(target_filename)

    ## Read columns or assume content
    assume_cols=['Acct #','Eff Date','Posted Date','Description','Hidden','Check #','Debit','Credit','Balance']
    assume_balance_cols=['Account','End Date','Expected Starting Balance','Expected Ending Balance','Calculated Ending Balance','Difference','Total Inflows','Total Outflows']

    entries=Excel.load_sheet('draft_entries')

    ## Output data
    ptr_row=2   #Yes, ACTUALLY second row (no row 0)
    ptr_col=1

    #for entry,target_cols in iter_source_data(target_cols=assume_cols):
    for entry,target_cols in iter_case_source_data(case_id,target_cols=assume_cols):
        ptr_col=1
        for col in assume_cols:
            entries.cell(row=ptr_row,column=ptr_col).value=entry[col]
            ptr_col+=1
        ptr_row+=1


    Excel.save()

    print ("[info] done excel saved to: "+str(target_filename))
    return target_filename,case_filename

def PREPARE_excel(case_id):
    version="1"
    case_filename='excel_case-'+str(case_id)+"-"+str(version)+'.xlsx'

    print ("[info] running excel for case: "+str(case_id))

    Excel=Excel_Wrap()

    excel_dir=LOCAL_PATH+"excel_files/"
    excel_case_dir=excel_dir+case_id+"/"
    # Create dir if not exists
    if not os.path.exists(excel_case_dir):
        os.makedirs(excel_case_dir)
    template_filename=excel_dir+"raw_template1.xlsx"
    target_filename=excel_case_dir+case_filename

    Excel.create_copy_from_to(template_filename,target_filename)

    print ("1)  Load template")

    Excel.open(target_filename)

    ## Read columns or assume content
    assume_balance_cols=['Account','End Date','Expected Starting Balance','Expected Ending Balance','Calculated Ending Balance','Difference','Total Inflows','Total Outflows']

    entries=Excel.load_sheet('draft_entries')

    return entries,Excel,target_filename

def dev_generate_excel_report_oct14(case_id=''):
    #0v3# Feb 17, 2024  Add demo checks output
    #0v2# Oct 14, 2024

    #** THIS IS 'OFFICIAL' ENTRYPOINT FOR GENERATING EXCEL  (called via pipeline + sim_user.py)
    #- recall, called from sim_user interface_dump_excel()

    print ("Integrate questions data maker")

    ## CALL DATA GENERATION
#    case_id='case_wells_fargo_small'
    erows=answer_excel(case_id=case_id)
    

    ## Check columns match
    assume_cols=['Acct #','Eff Date','Posted Date','Description','Hidden','Check #','Debit','Credit','Balance','Locate']
    
    ## Use same approach as above
    entries,Excel,target_filename=PREPARE_excel(case_id)

    ptr_row=2 #2 is second row
    for rowdict in erows:
#D#        print ("ROW DICT: "+str(rowdict))
        ptr_col=1
        row_headings=list(rowdict.keys())
        #^ should match assume_cols target excel template
        for col in assume_cols:
            cell_value=rowdict.get(col,'')
            
            if col=='Locate':
                # HYPERLINK!
                entries.cell(row=ptr_row,column=ptr_col).value='Locate pdf page'
                entries.cell(row=ptr_row,column=ptr_col).hyperlink=cell_value
                entries.cell(row=ptr_row,column=ptr_col).font=Font(color="0000FF", underline="single")
            else:
                entries.cell(row=ptr_row,column=ptr_col).value=cell_value
            ptr_col+=1

        ptr_row+=1
        
    ### DEMO CHECKS
    try:
        Excel=demo_generate_checks_tab(case_id=case_id,Excel=Excel, auto_save=False)
    except Exception as e:
        print ("[error] demo_generate_checks_tab: "+str(e))

    ## Standard save
    Excel.save()

    print ("[info] done excel saved to: "+str(target_filename))
    case_filename=os.path.basename(target_filename)
    return target_filename,case_filename


def call_generate_excel_report_oct14():
    case_id='pjplumbingA'
    case_id='65caaffb9b6ff316a779f525'
    
    dev_generate_excel_report_oct14(case_id=case_id)

    return


def demo_generate_checks_tab(case_id='65caaffb9b6ff316a779f525', Excel=None, auto_save=True):
    print ("DOING CHECKS REPORT FOR: "+str(case_id))

    ## SAMPLE EXCEL FILE START [ ] from above
    if not Excel:
        entries,Excel,target_filename=PREPARE_excel(case_id)
    
    ## Standard branches:
    b=['get_check_data']
    b+=['prepare_excel']

    if 'get_check_data' in b:
        
        ## IMPORT & SETTINGS
        from dev_check_report import get_sample_check_values  #Inner import for now
        
        ##################################
        ## Fetch checks data
        headings,records,meta=get_sample_check_values(case_id=case_id)
        
    
    if 'prepare_excel' in b:
        ## Recall Excel_Wrap is base library @
        #    from exceltemplate.excel_wrap import Excel_Wrap
        #    from openpyxl import Workbook
        #    from openpyxl import load_workbook
        #     So, direct access is Excel.wb


        ## CREATE NEW TAB FOR CHECKS OR USE TEMPLATE

        # If 'Checks' tab not exist then create in Excel.wb
        if 'Checks' not in Excel.wb.sheetnames:
            Excel.wb.create_sheet('Checks')
        checks=Excel.load_sheet('Checks')
        
        # Add exception clause at 1,1
        clause='*This list of extracted checks is not fully resolved (0v1)'
        checks.cell(row=1,column=1).value=clause
        
        # Add headings to row 3
        ptr_row=3
        ptr_col=1
        for h in headings:
            checks.cell(row=ptr_row,column=ptr_col).value=h
            #Bold cell
            checks.cell(row=ptr_row,column=ptr_col).font=Font(bold=True)
            ptr_col+=1 
            
        # Double width of all heading columns
        for i in range(1,len(headings)+1):
            checks.column_dimensions[chr(64+i)].width=20
        # Specific width for some columns if Payor=50, Payee 60, 
        for h in headings:
            if h=='Payor':
                checks.column_dimensions['E'].width=50
            if h=='Payee':
                checks.column_dimensions['F'].width=60
            
        # Add records
        ptr_row=4
        for record in records:
            ptr_col=1
            for field in record:
                
                ## Output value (start with exceptions)
                if field=='Image URL':
                    # HYPERLINK!
                    checks.cell(row=ptr_row,column=ptr_col).value='View check'
                    checks.cell(row=ptr_row,column=ptr_col).hyperlink=record[field]
                    checks.cell(row=ptr_row,column=ptr_col).font=Font(color="0000FF", underline="single")
                else:
                    checks.cell(row=ptr_row,column=ptr_col).value=record[field]

                ptr_col+=1
            ptr_row+=1
            
        if auto_save:
            filename=Excel.save()
            print ("Saved updated to: "+str(filename))


    print ("[ ] todo:  add test for current excel file dumper")
    return Excel



def doc_notes():
    print ("What is the current pipeline entrypoint for excel?")
    print ("   |--> w_ui/sim_user.py -> interface_dump_excel -> dev_ereport.py -> dev_generate_excel_report_oct14")

    return

if __name__=='__main__':
    branches=[]
    branches+=['dev_generate_excel_report_1']
    branches+=['dev_generate_excel_report_oct14']

    branches=['demo_generate_checks_tab']

    branches=['call_generate_excel_report_oct14']

    for b in branches:
        globals()[b]()
        

"""

Sheet:
draft_entries

Data:
Acct #	Eff Date	Posted Date	Description	Check #	Debit	Credit	Balance	
1010083000401	12/10/2018	12/10/2018	"12172 Highway 92 Woodstock GA 
0006791 ATM ID 6935B Card 9916"			$1,000.01	$6,951.88	Locate
1010083000401	12/07/2018	12/10/2018	"Redbox Dvd Rental 866-733-2693 
IL S468341595323292 Card 9916"		-$14.84		$6,937.04	Locate
1010083000401		12/11/2018	Albert Smith Recurring No Account Number on 12-11		-$150.00		$6,787.04	Locate


Sheet:
Balance Summary

Data:
Account	End Date	Expected Starting Balance	Expected Ending Balance	Calculated Ending Balance	Difference	Total Inflows	Total Outflows
1010083000401	01/08/2019	$5,951.87	$4,511.86	$4,511.86	$0.00	$10,051.32	$11,491.33
1010083000401	02/07/2019	$4,511.86	$6,251.80	$6,251.80	$0.00	$14,556.80	$12,816.86
1010083000401	03/07/2019	$6,251.80	$13,618.70	$13,618.70	$0.00	$18,427.09	$11,060.19



"""
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
